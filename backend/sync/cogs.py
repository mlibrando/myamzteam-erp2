"""COGS import and monthly cost-of-goods computation.

Reads COGS_Magical_Butter_1.xlsx (sheets US / CA / AU / UK) and:
  1. Upserts cogs_per_sku from the workbook.
  2. Joins quantity_shipped (sp_transaction_items) × cogs to produce monthly COGS.
  3. Writes pnl_monthly rows for cogs.cost_of_goods.
  4. Surfaces SKUs with sales but no COGS entry in cogs_missing_skus.

AU sheet has no Status column — all sheets are read by header name, not index.
Duplicate SKU within a sheet raises DuplicateSkuError before any DB write.

Usage:
    python -m sync.cogs [--marketplace US] [--log-level INFO]
"""

from __future__ import annotations

import argparse
import logging
import os
import pathlib
import sys
from collections import Counter
from decimal import Decimal

import openpyxl
import psycopg
from dotenv import load_dotenv

from .config import MARKETPLACE_ALIASES, MARKETPLACE_TO_SHEET, COGS_XLSX

log = logging.getLogger(__name__)


class DuplicateSkuError(ValueError):
    pass


def _load_sheet(sheet_name: str) -> list[dict]:
    """Return list of row dicts for one sheet, read by header name.

    Raises DuplicateSkuError if any SKU appears more than once.
    Formula cells are resolved via data_only=True (cached values).
    """
    wb = openpyxl.load_workbook(str(COGS_XLSX), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {COGS_XLSX.name}. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name: str) -> int | None:
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    sku_col = col("SKU")
    if sku_col is None:
        raise ValueError(f"Sheet '{sheet_name}' has no SKU column")

    col_asin     = col("ASIN")
    col_name     = col("Product Name")
    col_status   = col("Status")          # None for AU
    col_price    = col("Normal Product Price")
    col_cogs     = col("COGS")

    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, sku_col).value
        if not sku:
            continue
        # Excel numeric cells (e.g. barcode-style SKUs) come back as floats.
        # Convert whole-number floats to int before stringifying so "850251005008.0" → "850251005008".
        if isinstance(sku, float) and sku == int(sku):
            sku = int(sku)
        rows.append({
            "sku":          str(sku).strip(),
            "asin":         str(ws.cell(r, col_asin).value).strip()   if col_asin  and ws.cell(r, col_asin).value  else None,
            "product_name": str(ws.cell(r, col_name).value).strip()   if col_name  and ws.cell(r, col_name).value  else None,
            "status":       str(ws.cell(r, col_status).value).strip()  if col_status and ws.cell(r, col_status).value else None,
            "normal_price": Decimal(str(ws.cell(r, col_price).value)) if col_price and ws.cell(r, col_price).value is not None else None,
            "cogs":         Decimal(str(ws.cell(r, col_cogs).value))  if col_cogs  and ws.cell(r, col_cogs).value  is not None else None,
        })

    # Drop rows with no COGS value (e.g. discontinued products).
    rows = [r for r in rows if r["cogs"] is not None]

    counts = Counter(r["sku"] for r in rows)
    dups = [sku for sku, n in counts.items() if n > 1]
    if dups:
        raise DuplicateSkuError(
            f"Duplicate SKUs in sheet '{sheet_name}': {', '.join(sorted(dups))}. "
            "Fix the workbook before running the COGS import."
        )

    return rows


def import_cogs(conn: psycopg.Connection, marketplace_id: str, sheet_name: str) -> int:
    """Upsert cogs_per_sku from the workbook sheet. Returns the row count."""
    rows = _load_sheet(sheet_name)
    if not rows:
        log.warning("No COGS rows found in sheet '%s'", sheet_name)
        return 0

    with conn.cursor() as cur:
        cur.executemany(
            """
            INSERT INTO cogs_per_sku
                (marketplace_id, sku, asin, product_name, status, normal_price, cogs, imported_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, now())
            ON CONFLICT (marketplace_id, sku) DO UPDATE SET
                asin          = EXCLUDED.asin,
                product_name  = EXCLUDED.product_name,
                status        = EXCLUDED.status,
                normal_price  = EXCLUDED.normal_price,
                cogs          = EXCLUDED.cogs,
                imported_at   = now()
            """,
            [
                (marketplace_id, r["sku"], r["asin"], r["product_name"],
                 r["status"], r["normal_price"], r["cogs"])
                for r in rows
            ],
        )
    conn.commit()
    log.info("Imported %d COGS rows for marketplace %s", len(rows), marketplace_id)
    return len(rows)


def compute_monthly_cogs(conn: psycopg.Connection, marketplace_id: str) -> dict:
    """Compute monthly COGS and update pnl_monthly.

    Joins sp_transaction_items.quantity_shipped × cogs_per_sku.cogs.
    Surfaces unknown SKUs in cogs_missing_skus.

    Returns stats dict.
    """
    stats = {"months": 0, "missing_skus": 0}

    with conn.cursor() as cur:
        # Monthly COGS: sum(quantity_shipped × unit_cogs) joined to cogs_per_sku on SKU.
        # Only Shipment transactions count (not refunds, removals, reimbursements).
        cur.execute(
            """
            SELECT
                to_char(t.posted_at AT TIME ZONE 'UTC', 'YYYY-MM') AS year_month,
                SUM(i.quantity_shipped * c.cogs)                    AS cogs_amount
            FROM sp_transaction_items i
            JOIN sp_transactions t ON t.transaction_id = i.transaction_id
            JOIN cogs_per_sku c
              ON c.sku = i.sku AND c.marketplace_id = t.marketplace_id
            WHERE t.marketplace_id = %s
              AND t.is_deferred_release_event = false
              AND (t.raw_json->>'transactionType') = 'Shipment'
              AND i.quantity_shipped > 0
            GROUP BY 1
            ORDER BY 1
            """,
            (marketplace_id,),
        )
        monthly_cogs = {r[0]: Decimal(str(r[1])) for r in cur.fetchall()}

        # Find SKUs that have Shipment sales but no COGS row (unmatched by SKU).
        cur.execute(
            """
            SELECT DISTINCT i.sku, i.asin
            FROM sp_transaction_items i
            JOIN sp_transactions t ON t.transaction_id = i.transaction_id
            LEFT JOIN cogs_per_sku c
              ON c.sku = i.sku AND c.marketplace_id = t.marketplace_id
            WHERE t.marketplace_id = %s
              AND t.is_deferred_release_event = false
              AND (t.raw_json->>'transactionType') = 'Shipment'
              AND i.quantity_shipped > 0
              AND i.sku IS NOT NULL
              AND c.sku IS NULL
            """,
            (marketplace_id,),
        )
        missing_rows = cur.fetchall()

    # Upsert missing SKUs
    if missing_rows:
        stats["missing_skus"] = len(missing_rows)
        log.warning(
            "%d SKUs have sales but no COGS entry — writing to cogs_missing_skus",
            len(missing_rows),
        )
        with conn.cursor() as cur:
            for sku, asin in missing_rows:
                cur.execute(
                    """
                    INSERT INTO cogs_missing_skus (marketplace_id, sku, asin, first_seen, last_seen)
                    VALUES (%s, %s, %s, now(), now())
                    ON CONFLICT (marketplace_id, sku) DO UPDATE SET
                        last_seen = now()
                    """,
                    (marketplace_id, sku, asin),
                )
        conn.commit()

    # Write pnl_monthly rows for COGS. Sellerise's top-level `cog` field.
    pnl_rows = [
        (marketplace_id, ym, "cog", "Cost of goods sold",
         "cog", abs(amount), "USD")
        for ym, amount in sorted(monthly_cogs.items())
        if amount != 0
    ]

    with conn.cursor() as cur:
        cur.execute(
            "DELETE FROM pnl_monthly WHERE marketplace_id = %s AND bucket = 'cog'",
            (marketplace_id,),
        )
        if pnl_rows:
            cur.executemany(
                """
                INSERT INTO pnl_monthly
                    (marketplace_id, year_month, line_key, line_label, bucket, amount, currency, computed_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, now())
                """,
                pnl_rows,
            )
    conn.commit()

    stats["months"] = len(pnl_rows)
    log.info(
        "Wrote %d pnl_monthly COGS rows for %s (total %s missing SKUs)",
        len(pnl_rows), marketplace_id, stats["missing_skus"],
    )
    return stats


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="sync.cogs")
    parser.add_argument("--marketplace", default="US")
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    load_dotenv(pathlib.Path(__file__).resolve().parents[2] / ".env")
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("DATABASE_URL not set", file=sys.stderr)
        return 1

    marketplace_id = MARKETPLACE_ALIASES.get(args.marketplace.upper(), args.marketplace)
    sheet_name = MARKETPLACE_TO_SHEET.get(marketplace_id, args.marketplace.upper())

    with psycopg.connect(db_url) as conn:
        try:
            imported = import_cogs(conn, marketplace_id, sheet_name)
        except DuplicateSkuError as exc:
            log.error("COGS import aborted: %s", exc)
            return 1

        stats = compute_monthly_cogs(conn, marketplace_id)

    log.info(
        "COGS done: %d SKUs imported, %d COGS months, %d missing SKUs",
        imported, stats["months"], stats["missing_skus"],
    )
    return 0 if stats["missing_skus"] == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
