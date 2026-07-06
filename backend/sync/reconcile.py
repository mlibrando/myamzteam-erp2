"""Compare pnl_monthly against the Sellerise XLSX sheet.

Reads RAW_AMZ_US_SELLERISE.xlsx (sheet RAW_AMZ_US), compares each mapped row
to pnl_monthly, and writes a diff CSV to stdout (or --output <path>).

Rows with no line_key mapping (e.g. FBA reimbursement sub-types) are flagged as
SHEET_ONLY and included so nothing is silently ignored.

Usage:
    python -m sync.reconcile [--marketplace US] [--output diff.csv] [--log-level INFO]
"""

from __future__ import annotations

import argparse
import csv
import io
import logging
import os
import pathlib
import sys
from decimal import Decimal
from typing import Optional

import openpyxl
import psycopg
from dotenv import load_dotenv

from .config import MARKETPLACE_ALIASES

log = logging.getLogger(__name__)

# Absolute path from repo root so this file works regardless of cwd.
_REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
SELLERISE_XLSX = _REPO_ROOT / "reference" / "data" / "RAW_AMZ_US_SELLERISE.xlsx"
SHEET_NAME = "RAW_AMZ_US"

# Row number (1-based) in the sheet → (line_key_in_pnl_monthly | None, description)
#
# None means the row exists in the sheet but has no API-level equivalent.
# Section-header rows (sums of sub-rows) are excluded — compare at sub-row level only.
#
# Multiple rows may share the same line_key: they are summed before comparing to pnl.
# Use case: the sheet splits settled vs pending (e.g. rows 14 + 15 = fba.per_unit)
# but the SP-API includes both in a single line_key.
ROW_MAP: dict[int, tuple[Optional[str], str]] = {
    # ── Sales ────────────────────────────────────────────────────────────────
    3:  ("sales.product_sales",         "Product sales"),
    4:  ("sales.tax",                   "Tax (sales)"),
    5:  ("sales.promotion",             "Promotion"),
    6:  ("sales.shipping_charge",       "Shipping charge"),
    7:  ("sales.shipping_tax",          "Shipping tax"),
    8:  ("sales.gift_wrap",             "Gift wrap"),
    9:  ("sales.gift_wrap_tax",         "Gift wrap tax"),
    10: (None,                          "Shipping (settlement-period edge) — not separately tracked"),
    # ── Cost of Goods ─────────────────────────────────────────────────────────
    11: ("cogs.cost_of_goods",           "Cost of goods sold"),
    # ── Taxes ────────────────────────────────────────────────────────────────
    12: ("taxes.mft",                   "Taxes (MFT)"),
    # ── FBA fees: rows 14+15 sum to fba.per_unit ────────────────────────────
    # Row 15 (pending) is the deferred portion not yet settled at month-end;
    # our DEFERRED transactions include it — map both to the same line_key.
    14: ("fba.per_unit",                "Fba per unit fulfillment fee (settled)"),
    15: ("fba.per_unit",                "Fba fees (pending — same line as settled)"),
    # ── Referral fees: rows 17+20 sum to referral.commission ────────────────
    17: ("referral.commission",         "Commission (settled)"),
    18: ("referral.shipping_chargeback","Shipping chargeback"),
    19: ("referral.giftwrap_chargeback","Giftwrap chargeback"),
    20: ("referral.commission",         "Referral fee (pending — same line as settled)"),
    21: (None,                          "Poa service fee — not in Finances API"),
    22: (None,                          "Po a per unit fulfillment fee — not in Finances API"),
    # ── Storage fees ─────────────────────────────────────────────────────────
    23: ("storage.monthly",             "Storage fees"),
    # ── Advertising (Phase 4 — SP billed total only for now) ────────────────
    24: ("advertising.advertising_fee", "Advertising expenses (SP-API billed total; per-type is Phase 4)"),
    25: (None,                          "Sponsored brands — Phase 4 (Ads API)"),
    26: (None,                          "Sponsored display — Phase 4 (Ads API)"),
    27: (None,                          "Sponsored videos — Phase 4 (Ads API)"),
    28: (None,                          "Sponsored television — Phase 4 (Ads API)"),
    29: (None,                          "Sponsored products — Phase 4 (Ads API)"),
    # ── Refunds ──────────────────────────────────────────────────────────────
    31: ("refunds.tax",                 "Tax (refunds)"),
    32: ("refunds.tax_withheld",        "Tax withheld"),
    33: ("refunds.product_sales",       "Product sales (refunds)"),
    34: ("refunds.commission",          "Commission (refunds)"),
    35: ("refunds.refund_commission",   "Refund commission"),
    36: ("refunds.promotion",           "Promotion (refunds)"),
    37: ("refunds.shipping_charge",     "Shipping charge (refunds)"),
    38: ("refunds.shipping_chargeback", "Shipping chargeback (refunds)"),
    39: ("refunds.shipping_tax",        "Shipping tax (refunds)"),
    40: ("refunds.restocking_fee",      "Restocking fee"),
    41: ("refunds.gift_wrap",           "Gift wrap (refunds)"),
    42: ("refunds.gift_wrap_tax",       "Gift wrap tax (refunds)"),
    43: ("refunds.giftwrap_chargeback", "Giftwrap chargeback (refunds)"),
    44: ("refunds.goodwill",            "Goodwill"),
    # ── Other AMZ transactions ────────────────────────────────────────────────
    46: ("other_amz.inbound_transport",        "Inbound transportation fee"),
    # Rows 47, 51, 54, 61, 62, 65 all represent FBA reimbursements; the SP-API
    # aggregates them into one breakdown_type (FBAInventoryReimbursement).
    # Map all to reversal_reimbursement so the comparison sums the sheet sub-rows.
    47: ("other_amz.reversal_reimbursement",   "Reversal reimbursement"),
    48: ("other_amz.amazon_fees",              "Amazon fees"),
    49: ("other_amz.inbound_placement",        "Fba inbound placement service fee"),
    50: ("other_amz.premium_services",         "Premium services fee"),
    51: ("other_amz.reversal_reimbursement",   "Missing from inbound (FBA reimb sub-type — aggregated)"),
    52: ("other_amz.subscription",             "Subscription fee"),
    53: ("other_amz.compensated_clawback",     "Compensated clawback"),
    54: ("other_amz.reversal_reimbursement",   "Warehouse damage (FBA reimb sub-type — aggregated)"),
    55: (None,                                 "Fba fees (misc adj) — not in Finances API"),
    56: ("other_amz.disposal_complete",        "Disposal complete"),
    57: ("other_amz.removal_complete",         "Removal complete"),
    58: ("other_amz.compensated_clawback",     "Missing from inbound clawback (aggregated with compensated clawback)"),
    59: ("other_amz.storage_renewal",          "Storage renewal billing"),
    60: ("other_amz.inbound_transport",        "Inbound transportation fee - adjustment (same line)"),
    61: ("other_amz.reversal_reimbursement",   "Free replacement refund items (FBA reimb sub-type — aggregated)"),
    62: ("other_amz.reversal_reimbursement",   "Warehouse lost (FBA reimb sub-type — aggregated)"),
    63: ("other_amz.retrocharge",              "Retrocharge"),
    64: ("other_amz.liquidations",             "Liquidations"),
    65: ("other_amz.reversal_reimbursement",   "Removal order lost (FBA reimb sub-type — aggregated)"),
    66: ("other_amz.retrocharge_reversal",     "Retrocharge reversal"),
    67: (None,                                 "Amazon fees (reimb adj) — not separately tracked"),
    68: (None,                                 "Product charges (reimb adj) — not separately tracked"),
    69: (None,                                 "Fba fees (reimb adj) — not separately tracked"),
    70: (None,                                 "Liquidation adjustments — not separately tracked"),
    71: ("other_amz.fee_adjustment",           "Fee adjustment"),
}

DELTA_THRESHOLD = Decimal("0.01")


def _read_sheet(path: pathlib.Path) -> tuple[list[str], dict[int, dict[str, Decimal]]]:
    """Return (month_headers, {row_num: {month: amount}}) from the Sellerise XLSX."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[SHEET_NAME]

    # Row 1: "Parameter/Date", "Jan 2026", "Feb 2026", …
    headers = []
    col_to_month: dict[int, str] = {}
    for col in range(2, ws.max_column + 1):
        v = ws.cell(1, col).value
        if not v:
            break
        col_to_month[col] = str(v)
        headers.append(str(v))

    data: dict[int, dict[str, Decimal]] = {}
    for row_num, _ in ROW_MAP.items():
        row_vals: dict[str, Decimal] = {}
        for col, month in col_to_month.items():
            v = ws.cell(row_num, col).value
            if isinstance(v, (int, float)) and v != 0:
                row_vals[month] = Decimal(str(v))
        data[row_num] = row_vals

    return headers, data


def _read_pnl(conn: psycopg.Connection, marketplace_id: str) -> dict[str, dict[str, Decimal]]:
    """Return {line_key: {year_month: amount}} from pnl_monthly."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT line_key, year_month, amount FROM pnl_monthly WHERE marketplace_id = %s",
            (marketplace_id,),
        )
        rows = cur.fetchall()

    result: dict[str, dict[str, Decimal]] = {}
    for line_key, year_month, amount in rows:
        result.setdefault(line_key, {})[year_month] = Decimal(str(amount))
    return result


def _ym_to_sheet_header(year_month: str) -> str:
    """Convert '2026-01' → 'Jan 2026'."""
    import datetime
    dt = datetime.date(int(year_month[:4]), int(year_month[5:7]), 1)
    return dt.strftime("%b %Y")


def reconcile(
    conn: psycopg.Connection,
    marketplace_id: str,
    sheet_path: pathlib.Path = SELLERISE_XLSX,
) -> list[dict]:
    """Return a list of diff rows: {row_num, description, line_key, month, sheet, pnl, delta, status}.

    Sheet rows that share a line_key (e.g. rows 14+15 → fba.per_unit) are summed
    before comparison so the diff reflects the single pnl_monthly value correctly.
    """
    import datetime

    headers, sheet_data = _read_sheet(sheet_path)
    pnl_data = _read_pnl(conn, marketplace_id)

    # First row_num and description for each line_key (used in output).
    key_meta: dict[str, tuple[int, str]] = {}
    for row_num, (line_key, description) in ROW_MAP.items():
        if line_key is not None and line_key not in key_meta:
            key_meta[line_key] = (row_num, description)

    # ── 1. Sum sheet values per (line_key, ym) ───────────────────────────────
    sheet_by_key: dict[tuple[str, str], Decimal] = {}
    sheet_only_rows: list[dict] = []

    for row_num, (line_key, description) in ROW_MAP.items():
        sheet_row = sheet_data.get(row_num, {})
        for sheet_month, sheet_val in sheet_row.items():
            try:
                dt = datetime.datetime.strptime(sheet_month, "%b %Y")
                ym = dt.strftime("%Y-%m")
            except ValueError:
                continue

            if line_key is None:
                if abs(sheet_val) >= DELTA_THRESHOLD:
                    sheet_only_rows.append({
                        "row_num": row_num,
                        "description": description,
                        "line_key": "",
                        "year_month": ym,
                        "sheet": float(sheet_val),
                        "pnl": 0.0,
                        "delta": 0.0,
                        "status": "SHEET_ONLY",
                    })
            else:
                key = (line_key, ym)
                sheet_by_key[key] = sheet_by_key.get(key, Decimal("0")) + sheet_val

    # ── 2. Union of (line_key, ym) pairs from both sources ───────────────────
    all_keys: set[tuple[str, str]] = set(sheet_by_key.keys())
    for lk, ym_dict in pnl_data.items():
        for ym in ym_dict:
            all_keys.add((lk, ym))

    # ── 3. Compare ────────────────────────────────────────────────────────────
    diffs: list[dict] = []
    for line_key, ym in sorted(all_keys):
        sheet_val = sheet_by_key.get((line_key, ym), Decimal("0"))
        pnl_val = pnl_data.get(line_key, {}).get(ym, Decimal("0"))
        delta = pnl_val - sheet_val

        status = "OK" if abs(delta) < DELTA_THRESHOLD else "DIFF"
        row_num, description = key_meta.get(line_key, (0, line_key))

        diffs.append({
            "row_num": row_num,
            "description": description,
            "line_key": line_key,
            "year_month": ym,
            "sheet": float(sheet_val),
            "pnl": float(pnl_val),
            "delta": float(delta),
            "status": status,
        })

    diffs.extend(sheet_only_rows)
    return diffs


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="sync.reconcile")
    parser.add_argument("--marketplace", default="US")
    parser.add_argument("--output", help="CSV output path (default: stdout)")
    parser.add_argument("--all", action="store_true", help="Include OK rows (default: DIFF and SHEET_ONLY only)")
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    load_dotenv(_REPO_ROOT / ".env")
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("DATABASE_URL not set", file=sys.stderr)
        return 1

    marketplace_id = MARKETPLACE_ALIASES.get(args.marketplace.upper(), args.marketplace)

    with psycopg.connect(db_url) as conn:
        diffs = reconcile(conn, marketplace_id)

    if not args.all:
        diffs = [d for d in diffs if d["status"] != "OK"]

    fieldnames = ["row_num", "description", "line_key", "year_month", "sheet", "pnl", "delta", "status"]
    out = open(args.output, "w", newline="") if args.output else sys.stdout
    writer = csv.DictWriter(out, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(diffs)
    if args.output:
        out.close()

    # Summary
    diff_rows   = [d for d in diffs if d["status"] == "DIFF"]
    ok_rows     = [d for d in diffs if d["status"] == "OK"]
    sheet_only  = [d for d in diffs if d["status"] == "SHEET_ONLY"]

    log.info(
        "Reconcile %s: %d DIFF, %d OK, %d SHEET_ONLY",
        marketplace_id, len(diff_rows), len(ok_rows), len(sheet_only),
    )

    if diff_rows:
        total_delta = sum(abs(d["delta"]) for d in diff_rows)
        log.warning("Total absolute delta across all DIFF rows: $%.2f", total_delta)

    return 0 if not diff_rows else 1


if __name__ == "__main__":
    raise SystemExit(main())
